#!/usr/bin/env python3
"""Читает таблицу Троечки VIII Octobearfest в стенограмму (docs/replay-transcript.md).

Лист протоколов даёт по бою две строки — по строке на команду, восемнадцать
клеток «сколько из троих ответили верно», сумму, место и рейтинговый балл.
Кресла лист не пишет: в клетке счёт, а не кто именно сказал. Поэтому марки в
стенограмме — цифры 0..3 по вопросу, а кресла синтезирует кодек (как перестрелку
в ЭК: сумма верна, состав придуман, потому что лист хранит только её).

    uv run --with openpyxl python scripts/troika/read-troika-sheets.py troika.xlsx
"""

import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "studchr"))
from transcript import table  # noqa: E402

QUESTIONS = 3

# Блок за блоком: лист с таблицами групп, лист с протоколами, и сколько групп на
# листе. Второй этап держит по две группы на листе, третий — по одной.
BLOCKS = [
    ("s1", [(f"Группа {g}", f"Группа {g} — протоколы", [g]) for g in "ABCDEFGH"]),
    ("s2", [("Группы I и K", "Группы I и K — протоколы", ["I", "K"]),
            ("Группы J и L", "Группы J и L — протоколы", ["J", "L"])]),
    ("s3", [("Группа M", "Группа M — протоколы", ["M"]),
            ("Группа N", "Группа N — протоколы", ["N"])]),
]
# Группы в порядке, в котором их раскрывает схема: s1-g1 — это группа A.
GROUP_ORDER = {"s1": list("ABCDEFGH"), "s2": ["I", "J", "K", "L"], "s3": ["M", "N"]}


def cell(ws, row, col):
    return ws.cell(row=row, column=col).value


def find(ws, text, rows=range(1, 40), cols=range(1, 60)):
    """Первая клетка с таким текстом — заголовки стоят в разных местах листов."""
    for r in rows:
        for c in cols:
            if str(cell(ws, r, c) or "").strip() == text:
                return r, c
    return None


def group_table(ws, letter):
    """Строки таблицы группы: id, название, забито, пропущено, очки, место."""
    head = find(ws, f"Таблица группы {letter}") or (3, 1)
    hrow = head[0] + (2 if find(ws, f"Таблица группы {letter}") else 0)
    # Заголовок таблицы — строка, где стоит «Команда» под её названием.
    while str(cell(ws, hrow, 2) or "").strip() != "Команда":
        hrow += 1
    columns = {}
    for c in range(1, 40):
        name = str(cell(ws, hrow, c) or "").strip()
        if name in ("Заб.", "Проп.", "Очки", "Место"):
            columns[name] = c
    rows = []
    r = hrow + 1
    while True:
        ident = str(cell(ws, r, 1) or "").strip()
        name = str(cell(ws, r, 2) or "").strip()
        if not ident or not name:
            break
        rows.append({
            "id": ident,
            "name": name,
            "scored": int(cell(ws, r, columns["Заб."]) or 0),
            "conceded": int(cell(ws, r, columns["Проп."]) or 0),
            "rating": float(cell(ws, r, columns["Очки"]) or 0),
            "place": float(cell(ws, r, columns["Место"]) or 0),
        })
        r += 1
    return rows


def bouts(ws, names):
    """Бои листа протоколов, по порядку: две команды с клетками, суммой и местом.

    Лист второго этапа держит бои двух групп подряд, поэтому здесь читаются все,
    а по группам их разводит вызывающий — по тому, чьи это команды.
    """
    out = []
    for r in range(1, ws.max_row + 1):
        label = str(cell(ws, r, 1) or "").strip()
        if not label.startswith("Бой "):
            continue
        # Шапка боя: строка «№ | ID | Команда | Сумма | Тема 1 …» ниже метки.
        head = r
        while head < ws.max_row and str(cell(ws, head, 1) or "").strip() != "№":
            head += 1
        if str(cell(ws, head, 1) or "").strip() != "№":
            continue
        # Сколько тем играет бой, говорит его собственная строка номиналов:
        # третий групповой этап играет восемь, первые два — шесть.
        cells = 0
        while cell(ws, head + 1, 5 + cells) not in (None, ""):
            cells += 1
        first = head + 2  # шапка, строка номиналов, затем команды
        seats = []
        for k in range(2):
            row = first + k
            ident = str(cell(ws, row, 2) or "").strip()
            if not ident or ident not in names:
                break
            marks = [cell(ws, row, 5 + i) for i in range(cells)]
            seats.append({
                "id": ident,
                "name": names[ident],
                "marks": [int(m) if m else 0 for m in marks],
                "total": int(cell(ws, row, 5 + cells) or 0),
                "place": float(cell(ws, row, 6 + cells) or 0),
            })
        if len(seats) == 2:
            out.append(seats)
    return out


def marks_field(marks):
    """Клетки по три на тему: «131 ..1 …», точка — вопрос не взял никто."""
    themes = []
    for t in range(0, len(marks), QUESTIONS):
        themes.append("".join(str(m) if m else "." for m in marks[t:t + QUESTIONS]))
    return " ".join(themes)


def main(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    # Финал и матч за 3-е место: по три боя, лист зовёт их O1..O3 и P1..P3.
    # В схеме это один блок, и бронза в нём стоит первой — её играют раньше, —
    # так что по координате s4/r1/w1 бои идут m1..m3 бронза, m4..m6 финал.
    FINALS = [("P", 1), ("O", 4)]
    out = ["# Троечка VIII Octobearfest — собрано read-troika-sheets.py из таблицы турнира",
           "[game]", "type: troika", "title: Троечка", "scheme: troika.dsl", ""]

    # Ростер — команды первого этапа, по номеру посева: группа A даёт 1..6 и так далее.
    names, roster = {}, []
    for _, blocks in BLOCKS:
        for sheet, _, letters in blocks:
            for letter in letters:
                for row in group_table(wb[sheet], letter):
                    names[row["id"]] = row["name"]
    # Посев. Лист называет команду «3A» — третья в группе A, — а схема раздаёт
    # ранги змейкой: первая восьмёрка по группам слева направо, вторая справа
    # налево. Значит ростер нумеруется обратной змейкой, чтобы восьмёрка групп
    # сошлась с листом. Сам посев здесь вход, а не проверка: его считал
    # оргкомитет по двум другим дисциплинам феста (регламент 4.4.2).
    letters = GROUP_ORDER["s1"]
    groups = len(letters)
    by_rank = {}
    for letter in letters:
        group = letters.index(letter) + 1
        for band, row in enumerate(group_table(wb[f"Группа {letter}"], letter)):
            rank = band * groups + (group if band % 2 == 0 else groups + 1 - group)
            by_rank[rank] = row["name"]
    for rank in sorted(by_rank):
        roster.append(f"{rank:>2} | {by_rank[rank]}")
    out += ["[roster]"] + roster

    for block, sheets in BLOCKS:
        for sheet, protocols, letters in sheets:
            sheet_names = {}
            for letter in letters:
                for row in group_table(wb[sheet], letter):
                    sheet_names[row["id"]] = row["name"]
            played = bouts(wb[protocols], sheet_names)
            for letter in letters:
                group = GROUP_ORDER[block].index(letter) + 1
                rows = group_table(wb[sheet], letter)
                mine = {row["id"] for row in rows}
                ours = [seats for seats in played if seats[0]["id"] in mine]
                # Группа из N играет круги по N/2 боёв: круг — это раунд,
                # а бои круга идут заходами, по одному на площадку.
                per_round = len(rows) // 2
                for index, seats in enumerate(ours):
                    coord = f"{block}/g{group}/r{index // per_round + 1}/w{index % per_round + 1}/m1"
                    out += ["", f"[{coord}] жребий"]
                    width = max(len(seat["name"]) for seat in seats)
                    for seat in seats:
                        out.append(f"{seat['name']:<{width}} | {marks_field(seat['marks'])} "
                                   f"| {seat['total']:>2} | {seat['place']:g}")
                table(out, f"{block}/g{group}", [(row["place"], row["name"]) for row in rows])

    ws = wb["Финалы — протоколы"]
    finals = {}
    for r in range(1, ws.max_row + 1):
        label = str(cell(ws, r, 1) or "").strip()
        if len(label) != 2 or label[0] not in "OP" or not label[1].isdigit():
            continue
        head = r
        while head < ws.max_row and str(cell(ws, head, 1) or "").strip() != "№":
            head += 1
        cells = 0
        while cell(ws, head + 1, 5 + cells) not in (None, ""):
            cells += 1
        seats = []
        for k in range(2):
            row = head + 2 + k
            marks = [cell(ws, row, 5 + i) for i in range(cells)]
            seats.append({
                "name": str(cell(ws, row, 3) or "").strip(),
                "marks": [int(m) if m else 0 for m in marks],
                "total": int(cell(ws, row, 5 + cells) or 0),
                "place": float(cell(ws, row, 6 + cells) or 0),
            })
        finals[label] = seats
    for prefix, first in FINALS:
        for k in (1, 2, 3):
            seats = finals.get(f"{prefix}{k}")
            if not seats:
                continue
            out += ["", f"[s4/r1/w1/m{first + k - 1}]"]
            width = max(len(seat["name"]) for seat in seats)
            for seat in seats:
                # Лист зовёт команду по полному имени, ростер — тоже.
                out.append(f"{seat['name']:<{width}} | {marks_field(seat['marks'])} "
                           f"| {seat['total']:>2} | {seat['place']:g}")

    Path("testdata/octobearfest2025").mkdir(parents=True, exist_ok=True)
    Path("testdata/octobearfest2025/troika.transcript").write_text("\n".join(out) + "\n", encoding="utf-8")
    print(f"боёв: {sum(1 for line in out if line.startswith('['))}")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "troika.xlsx")
