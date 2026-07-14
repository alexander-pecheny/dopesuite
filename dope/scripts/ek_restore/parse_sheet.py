#!/usr/bin/env python3
"""Parse the СтудЧР-2026 ЭК Google-Sheet export into a structured JSON.

Each protocol tab (1/16, 1/8, 1/4, 1/2, Финал) holds one or more match blocks:
  header row:  c1="Бой X", then per-theme columns [10,20,30,40,50, "Тn", sep]
  team blocks: 2 rows each
    row A: c1=team name, c2=Σ, c3=М (place), c4=hidden name,
           then per theme: player-name in the "10" column, theme subtotal in "Тn"
    row B: per theme: the 5 question marks in the [10..50] columns
The mark vocabulary mirrors the app's normalizeMark: 1/q/й/+ -> right, -1/w/ц/-/− -> wrong.
"""
import json
import sys

import openpyxl

XLSX = "/tmp/ek_sheet.xlsx"
QUESTION_VALUES = [10, 20, 30, 40, 50]

# tab name -> list of expected match codes (for sanity), and stage code in DB
TAB_STAGES = [
    ("116", "r16"),   # A-F (run1) + G-L (run2)
    ("18", "r8"),     # M-R
    ("14", "r4"),     # S-V (3 teams each)
    ("12", "r2"),     # W-X
    ("Финал", "final"),  # Y
]


def norm_mark(v):
    """Return 'right'/'wrong'/'' for a sheet cell value."""
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        if v == 1:
            return "right"
        if v == -1:
            return "wrong"
        return ""
    s = str(v).strip().lower()
    if s in ("right", "q", "й", "1", "1.0", "+"):
        return "right"
    if s in ("wrong", "w", "ц", "-1", "-1.0", "-", "−1", "−"):
        return "wrong"
    return ""


def find_theme_cols(ws, hrow, maxcol):
    """Theme start columns: a cell ==10 followed by 20,30,40,50."""
    cols = []
    for c in range(1, maxcol - 3):
        vals = [ws.cell(row=hrow, column=c + k).value for k in range(5)]
        if vals == [10.0, 20.0, 30.0, 40.0, 50.0] or vals == [10, 20, 30, 40, 50]:
            cols.append(c)
    return cols


def parse_tab(ws):
    maxrow, maxcol = ws.max_row, ws.max_column
    matches = []
    r = 1
    while r <= maxrow:
        c1 = ws.cell(row=r, column=1).value
        if isinstance(c1, str) and c1.strip().startswith("Бой "):
            code = c1.strip().split()[1]
            theme_cols = find_theme_cols(ws, r, maxcol)
            # team rows follow until a blank c1 / next "Бой" header
            teams = []
            rr = r + 1
            while rr <= maxrow:
                name = ws.cell(row=rr, column=1).value
                if name is None or str(name).strip() == "":
                    break
                if str(name).strip().startswith("Бой "):
                    break
                name = str(name).strip()
                sigma = ws.cell(row=rr, column=2).value
                place = ws.cell(row=rr, column=3).value
                themes = []
                for ti, tc in enumerate(theme_cols):
                    player = ws.cell(row=rr, column=tc).value
                    player = str(player).strip() if player not in (None, "") else ""
                    marks = [norm_mark(ws.cell(row=rr + 1, column=tc + k).value) for k in range(5)]
                    themes.append({"theme_index": ti, "player": player, "marks": marks})
                teams.append({
                    "name": name,
                    "sigma": sigma,
                    "place": place,
                    "themes": themes,
                })
                rr += 2
            matches.append({"code": code, "theme_cols": theme_cols, "teams": teams})
            r = rr
        else:
            r += 1
    return matches


def compute_total(themes):
    total = 0
    for th in themes:
        for k, m in enumerate(th["marks"]):
            if m == "right":
                total += QUESTION_VALUES[k]
            elif m == "wrong":
                total -= QUESTION_VALUES[k]
    return total


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    out = {"stages": []}
    mismatches = []
    for tab, stage in TAB_STAGES:
        ws = wb[tab]
        matches = parse_tab(ws)
        out["stages"].append({"tab": tab, "stage": stage, "matches": matches})
        for m in matches:
            for t in m["teams"]:
                ct = compute_total(t["themes"])
                sig = t["sigma"]
                sig_i = int(sig) if isinstance(sig, (int, float)) else None
                flag = "" if sig_i == ct else "  <<< MISMATCH"
                if sig_i != ct:
                    mismatches.append((tab, m["code"], t["name"], sig_i, ct))
                print(f"{tab:6} Бой {m['code']:2} {t['name'][:28]:28} place={str(t['place']):4} Σsheet={str(sig_i):5} Σcalc={ct:5}{flag}")
    with open("/tmp/ek_parsed.json", "w") as f:
        json.dump(out, f, ensure_ascii=False, indent=1)
    print()
    print(f"Wrote /tmp/ek_parsed.json")
    print(f"Total matches: {sum(len(s['matches']) for s in out['stages'])}")
    print(f"TOTAL mismatches: {len(mismatches)}")
    for mm in mismatches:
        print("  MISMATCH", mm)


if __name__ == "__main__":
    main()
