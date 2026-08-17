"""Reads СтудЧР's личная СИ workbook into the shape the transcript wants:
players in seed order, group membership, and every бой's marks, Σ and место.

The grid decoding — and the check that it is right — lives in sheetgrid.py,
shared with ТПШ."""
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
import sheetgrid

SRC = "sheets/sheet-1HOqiPxINFxW3NVu6QAOKuU8yOyXB3IwjGCrO6AqHtK4.xlsx"

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
players = []
for row in wb["Регистрация"].iter_rows(min_row=2, max_col=5, values_only=True):
    if row[0]:
        players.append({"name": str(row[0]).strip(), "seed": int(row[4]) if row[4] else None})

groups, header = {}, None
for row in wb["Группы"].iter_rows(max_col=24, values_only=True):
    cells = [str(c).strip() if c else "" for c in row]
    if any(c.startswith("Группа ") for c in cells):
        header = {i: c.split()[-1] for i, c in enumerate(cells) if c.startswith("Группа ")}
        continue
    if not header:
        continue
    for i, name in enumerate(cells):
        if i in header and name and not name.startswith("Пл."):
            groups.setdefault(header[i], []).append(name)

# «Группы A-B» and its siblings print each group twice: as entered, and sorted
# by Очки from column H on. The sorted one is the group's table.
tables = {}
for title in ["Группы A-B", "Группы C-D", "Группы E-F"]:
    letter = None
    for row in wb[title].iter_rows(max_col=8, values_only=True):
        cell = str(row[7]).strip() if row[7] is not None else ""
        if cell.startswith("Группа "):
            letter = cell.split()[-1]
            tables[letter] = []
        elif letter and cell and cell != "Игрок":
            tables[letter].append(cell)
for letter, names in groups.items():
    if sorted(tables.get(letter, [])) != sorted(names):
        sys.exit(f"таблица группы {letter} не сходится с её составом: {tables.get(letter)} vs {names}")

rounds = {}
for title in ["Круг 1 (протоколы)", "Круг 2 (протоколы)", "Круг 3 (протоколы)", "Круг 4 (протоколы)"]:
    rounds[title] = sheetgrid.read_bouts(wb[title])
playoff = sheetgrid.read_bouts(wb["Плей-офф (протоколы)"])
# The grand final sits on its own sheet because it is played over twelve themes
# where the rest of the play-off has eight, so its grid is a different width.
playoff += sheetgrid.read_bouts(wb["Грандфинал (протокол)"])

# «Статистика»: Игрок | Счёт | Без − (Σ+) | Бои. Recomputed from the decoded
# marks and held to the tab — with one reviewed exception, where the tab does
# not add up against its own protocols (its «Без −» counts one right «20» more
# than they contain; Σ agrees both ways). That line rides into the transcript
# and is silenced there by an override.
KNOWN_BAD = {"Станислав Хамидулин"}
stats = []
for row in wb["Статистика"].iter_rows(values_only=True):
    if not row or not row[0] or str(row[0]).strip() == "Игрок" or row[1] is None:
        continue
    stats.append({"player": str(row[0]).strip(), "sum": int(row[1]),
                  "plus": int(row[2]), "bouts": int(row[3])})
wb.close()

computed = {}
for bouts in list(rounds.values()) + [playoff]:
    for bout in bouts:
        for seat in bout["players"]:
            entry = computed.setdefault(seat["name"], [0, 0, 0])
            entry[2] += 1
            for theme in seat["themes"]:
                for k, m in enumerate(theme):
                    if m == "right":
                        entry[0] += 10 * (k + 1)
                        entry[1] += 10 * (k + 1)
                    elif m == "wrong":
                        entry[0] -= 10 * (k + 1)
sheet_by = {s["player"]: [s["sum"], s["plus"], s["bouts"]] for s in stats}
bad = [name for name in set(computed) | set(sheet_by)
       if computed.get(name) != sheet_by.get(name) and name not in KNOWN_BAD]
if bad:
    sys.exit(f"статистика не сходится с протоколами: {sorted(bad)}")

out = {"players": players, "groups": groups, "tables": tables, "rounds": rounds, "playoff": playoff, "stats": stats}
json.dump(out, open("si-data.json", "w"), ensure_ascii=False)
print("players", len(players), "groups", {g: len(v) for g, v in groups.items()})
for title, bouts in rounds.items():
    print(title, len(bouts), "боёв, первый:", bouts[0]["code"], [p["name"] for p in bouts[0]["players"]])
print("плей-офф боёв:", len(playoff))
print("тем, где сумма не сошлась:", len(sheetgrid.MISREADS))
for row in sheetgrid.MISREADS[:5]:
    print("   ", row)
