"""Reads СтудЧР's ТПШ workbook: a written round of 91, then a bracket of 24.

The grids are личная СИ's shape, so the decoding is sheetgrid's. The written
round prints no место — «Итоги отбора» carries the ranking instead, and that is
what the replay holds dope to."""
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
import sheetgrid

SRC = "sheets/sheet-1skqr_3z7neSCh-u_wDPZ_l1zYcYnVAPe4LST3KgSuMI.xlsx"

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

# «Итоги отбора» ranks the written round by Σ, then Σ+, then how many 50s, 40s,
# 30s and 20s each player took — the chain the scheme's sorting spells out.
standing = {}
for row in wb["Итоги отбора"].iter_rows(min_row=2, max_col=2, values_only=True):
    if row[0] and row[1]:
        standing[str(row[1]).strip()] = int(row[0])

written = sheetgrid.read_bouts(wb["Письменный отбор"])
if len(written) != 1:
    sys.exit(f"письменный отбор — это один бой, а прочитано {len(written)}")
for player in written[0]["players"]:
    place = standing.get(player["name"])
    if place is None:
        sys.exit(f"{player['name']} писал отбор, но его нет в «Итогах отбора»")
    player["place"] = place

playoff = sheetgrid.read_bouts(wb["Плей-офф (протоколы)"])
registered = [str(row[0]).strip() for row in wb["Регистрация"].iter_rows(min_row=2, values_only=True) if row[0]]
wb.close()

json.dump({"players": registered, "written": written[0], "playoff": playoff},
          open("tpsh-data.json", "w"), ensure_ascii=False)
print("зарегистрировано", len(registered), "| писало отбор", len(written[0]["players"]))
print("боёв плей-офф:", len(playoff), [b["code"] for b in playoff])
print("тем, где сумма не сошлась:", len(sheetgrid.MISREADS))
for row in sheetgrid.MISREADS[:5]:
    print("   ", row)
