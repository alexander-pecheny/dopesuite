"""Reads СтудЧР's брейн (КИнСБФ) workbook: four protocol sheets, 132 бои.

A брейн protocol is laid out across the page rather than down it: бои sit side by
side in blocks seven columns wide, each headed by its code, the two teams and the
score, with the questions underneath and both teams' составы below those.

    A | Рыб'ending | 4 : 0 |   | Постпопс |
    1 | Корнеева   | 1     |   |          |
    2 |            |       | 0 | Алексеев |
      | Васильев   |       |   | Андреева |   <- составы start after a blank

A question can be marked on both sides: in брейн one team buzzes and misses, then
the other does. 1 is taken, 0 is missed, blank is untouched. The «П» rows are
перестрелка.
"""
import json
import os
import re
import sys

import openpyxl

SRC = "sheets/sheet-1M4_-FnE01tIvz9Hd5f2NJm_zE9tHJ7JJ-YHrBPrLxpE.xlsx"
SHEETS = ["1-й групповой этап (протоколы)", "DE (протоколы)",
          "2-й групповой этап (протоколы)", "3-й групповой этап (протоколы)",
          "Финальный этап (протоколы)"]
STRIDE, FIRST = 7, 1
SCORE = re.compile(r"^\d+\s*:\s*\d+$")


def text(cell):
    return str(cell).strip() if cell is not None else ""


def mark(cell):
    """1 taken, 0 missed, blank untouched. The sheet writes them as numbers."""
    if cell is None or text(cell) == "":
        return ""
    return "right" if float(cell) > 0 else "wrong"


def blocks_at(rows, top, group, title):
    """Every бой whose header is on row `top`, read down to the blank line.

    Blocks are found by their score cell rather than by counting columns: the
    group sheets start their first block at column 1 and the final sheet at
    column 0, and a score is the one cell no other row can be mistaken for."""
    header = rows[top]
    out = []
    for column, cell in enumerate(header):
        if column < 2 or not SCORE.match(text(cell)):
            continue
        base = column - 2
        if base + 5 >= len(header):
            continue
        left, right = text(header[base + 1]), text(header[base + 4])
        scored = [int(n) for n in text(header[base + 2]).split(":")]
        bout = {"group": group, "title": title, "code": text(header[base]),
                "teams": [{"name": left, "score": scored[0], "roster": []},
                          {"name": right, "score": scored[1], "roster": []}],
                "questions": []}
        for row in rows[top + 1:]:
            if base + 4 >= len(row):
                break
            label, p1, m1, m2, p2 = (text(row[base]), text(row[base + 1]), row[base + 2],
                                     row[base + 3], text(row[base + 4]))
            if label:  # a question row: «1», «2», … or «П»
                bout["questions"].append({
                    "tiebreak": label == "П",
                    "left": {"player": p1, "mark": mark(m1)},
                    "right": {"player": p2, "mark": mark(m2)}})
                continue
            if p1 or p2:  # the составы below the questions
                if p1:
                    bout["teams"][0]["roster"].append(p1)
                if p2:
                    bout["teams"][1]["roster"].append(p2)
                continue
            if bout["questions"]:
                break
        out.append(bout)
    return out


def read_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    out, group, title = [], "", ""
    for i, row in enumerate(rows):
        if any(SCORE.match(text(c)) for c in row):
            out += blocks_at(rows, i, group, title)
            continue
        for label in (text(c) for c in row if text(c)):
            if label.startswith("Группа "):
                group = label.split()[1].strip("(")
                break
            if not label.replace(".", "").isdigit():
                title = label
                break
    return out


def read_stats(ws):
    """«Статистика»: Игрок | Команда | Попытки | Верно | Неверно."""
    out = []
    for row in ws.iter_rows(values_only=True):
        if not row or not row[0] or text(row[0]) == "Игрок" or row[2] is None:
            continue
        out.append({"player": text(row[0]), "team": text(row[1]),
                    "attempts": int(row[2]), "right": int(row[3]), "wrong": int(row[4])})
    return out


def check_stats(stages, stats):
    """The tab's aggregates must equal what the decoded questions add up to.
    Перестрелки stay out: the tab does not count them."""
    computed = {}
    for bouts in stages.values():
        for b in bouts:
            for question in b["questions"]:
                if question["tiebreak"]:
                    continue
                for side, key in ((0, "left"), (1, "right")):
                    cell = question[key]
                    if not cell["player"] or not cell["mark"]:
                        continue
                    entry = computed.setdefault((cell["player"], b["teams"][side]["name"]), [0, 0, 0])
                    entry[0] += 1
                    entry[1 if cell["mark"] == "right" else 2] += 1
    sheet = {(s["player"], s["team"]): [s["attempts"], s["right"], s["wrong"]] for s in stats}
    bad = [key for key in set(computed) | set(sheet) if computed.get(key) != sheet.get(key)]
    if bad:
        sys.exit(f"статистика не сходится с протоколами: {sorted(bad)[:5]} и ещё {max(len(bad) - 5, 0)}")


if __name__ == "__main__":
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    stages = {name: read_sheet(wb[name]) for name in SHEETS}
    lineups = {}
    for row in wb["Составы"].iter_rows(values_only=True):
        cells = [text(c) for c in row]
        if cells and cells[0]:
            lineups.setdefault(cells[0], []).extend(c for c in cells[1:] if c)
    stats = read_stats(wb["Статистика"])
    wb.close()
    check_stats(stages, stats)
    json.dump({"stages": stages, "lineups": lineups, "stats": stats},
              open("brain-data.json", "w"), ensure_ascii=False)
    total = 0
    for name, bouts in stages.items():
        total += len(bouts)
        groups = sorted({b["group"] for b in bouts})
        print(f"{name}: {len(bouts)} боёв, групп {len(groups)} {groups}")
    print("всего боёв:", total, "| составов в реестре:", len(lineups))
    # The score the sheet printed must equal the questions it recorded, or the
    # decoding is wrong somewhere.
    bad = 0
    for bouts in stages.values():
        for b in bouts:
            for side, key in ((0, "left"), (1, "right")):
                took = sum(1 for q in b["questions"] if q[key]["mark"] == "right")
                if took != b["teams"][side]["score"]:
                    bad += 1
                    if bad < 6:
                        print("  счёт не сходится:", b["code"], b["teams"][side]["name"],
                              took, "vs", b["teams"][side]["score"])
    print("боёв, где счёт не сходится с вопросами:", bad)
