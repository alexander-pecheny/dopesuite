#!/usr/bin/env python3
"""Читает таблицу «Ассорти» в набор для теста (см. рядом README.md).

Две мини-игры, каждая со своим листом, и общая таблица, где каждая нормирована
к сотне и они сложены. Лист «Медиа-Эрудит» — темы по десять вопросов номиналами
10,10,20,20,30,30,40,40,50,50, клетка 1 / -1 / пусто; «Не только песни» — просто
столбцы по баллу.

    uv run --with openpyxl python scripts/multi/read-assorti-sheets.py assorti.xlsx
"""

import json
import sys
from pathlib import Path

import openpyxl

# Номиналы темы «Медиа-эрудита»: каждый по два вопроса.
MEDIA_NOMINALS = [10, 10, 20, 20, 30, 30, 40, 40, 50, 50]


def num(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def cell(ws, row, col):
    return ws.cell(row=row, column=col).value


def read_media(ws):
    """Темы по десять вопросов: клетка — верно, неверно или пусто."""
    themes = sum(1 for c in range(5, ws.max_column + 1)
                 if str(cell(ws, 2, c) or "").strip().startswith("Тема"))
    rows = []
    for r in range(4, ws.max_row + 1):
        name = cell(ws, r, 2)
        if not name:
            continue
        total, points = num(cell(ws, r, 3)), num(cell(ws, r, 4))
        declined = isinstance(cell(ws, r, 4), str)  # «Вне зачёта»
        if total is None:
            continue
        cells, col = [], 5
        for _ in range(themes):
            for nominal in MEDIA_NOMINALS:
                mark = num(cell(ws, r, col))
                cells.append(int(nominal * mark) if mark else 0)
                col += 1
            col += 2  # столбцы «Тема» и «Σ» после каждой темы
        rows.append({"name": str(name).strip(), "cells": cells,
                     "total": int(total), "points": points, "declined": declined})
    one = " ".join(f"{{-{n},0,{n}}}x2" for n in (10, 20, 30, 40, 50))
    return rows, " ".join([one] * themes), themes * len(MEDIA_NOMINALS)


def read_songs(ws):
    """Столбцы по баллу: клетка — взяли или нет."""
    width = sum(1 for c in range(5, ws.max_column + 1) if num(cell(ws, 1, c)) == 1)
    rows = []
    for r in range(4, ws.max_row + 1):
        name = cell(ws, r, 2)
        if not name:
            continue
        total, points = num(cell(ws, r, 3)), num(cell(ws, r, 4))
        declined = isinstance(cell(ws, r, 4), str)
        if total is None:
            continue
        rows.append({"name": str(name).strip(),
                     "cells": [1 if num(cell(ws, r, 5 + i)) else 0 for i in range(width)],
                     "total": int(total), "points": points, "declined": declined})
    return rows, f"{{0,1}}x{width}", width


def main(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    media, media_spec, media_width = read_media(wb["Медиа-Эрудит"])
    songs, songs_spec, songs_width = read_songs(wb["Не только песни"])

    names, order = {}, []
    for rows in (media, songs):
        for row in rows:
            if row["name"] not in names:
                names[row["name"]] = len(order) + 1
                order.append(row["name"])
    # Вне зачёта хотя бы в одной игре — вне зачёта: лист не считает такой
    # команде очков и не даёт ей задавать масштаб остальным.
    declined = sorted({row["name"] for rows in (media, songs) for row in rows if row["declined"]})

    overall = []
    ws = wb["Общая"]
    for r in range(3, ws.max_row + 1):
        name = cell(ws, r, 2)
        total = num(cell(ws, r, 3))
        if not name or total is None:
            continue
        overall.append({"place": num(cell(ws, r, 1)), "name": str(name).strip(), "total": total})

    def grid(rows, width):
        out = []
        for name in order:
            found = next((row for row in rows if row["name"] == name), None)
            out.append(found["cells"] if found else [0] * width)
        return out

    fixture = {
        "spec": f"Медиа-эрудит →0..100: {media_spec}\nНе только песни →0..100: {songs_spec}",
        "participants": [{"number": names[name], "name": name} for name in order],
        "declined": declined,
        "games": [grid(media, media_width), grid(songs, songs_width)],
        "minigameTotals": [{row["name"]: row["total"] for row in media},
                           {row["name"]: row["total"] for row in songs}],
        "minigamePoints": [{row["name"]: row["points"] for row in media if row["points"] is not None},
                           {row["name"]: row["points"] for row in songs if row["points"] is not None}],
        "overall": overall,
    }
    out = Path("testdata/assorti2025")
    out.mkdir(parents=True, exist_ok=True)
    (out / "assorti.json").write_text(json.dumps(fixture, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"команд {len(order)}, вне зачёта {len(declined)}, в общей {len(overall)}")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "assorti.xlsx")
